#!/usr/bin/env python3
"""Quem abre chamado no periodo noturno, a partir do export detalhado do BI.

Abertura = primeira hora em que o chamado aparece.
Noite estrita = 19h as 23h. Madrugada = 00h as 03h.
"""
import sys
from collections import Counter, defaultdict

import openpyxl

arq = sys.argv[1] if len(sys.argv) > 1 else "detalhada_2607.xlsx"
wb = openpyxl.load_workbook(arq, data_only=True)
ws = wb["Export"]

primeira, meta = {}, {}
perda_ch = defaultdict(float)

for r in range(2, ws.max_row + 1):
    regional = ws.cell(row=r, column=3).value
    loja = ws.cell(row=r, column=5).value
    data = ws.cell(row=r, column=6).value
    hora = ws.cell(row=r, column=7).value
    subcat = ws.cell(row=r, column=10).value
    chamados = ws.cell(row=r, column=11).value
    perda = float(ws.cell(row=r, column=12).value or 0)
    if regional is None or data is None or hora is None:
        continue
    h = int(hora)
    for c in str(chamados or "").split(","):
        c = c.strip()
        if not c:
            continue
        perda_ch[c] += perda / max(len(str(chamados).split(",")), 1)
        chave = (data.date(), h)
        if c not in primeira or chave < primeira[c]:
            primeira[c] = chave
            meta[c] = (regional, loja or "", subcat or "")

NOITE = set(range(19, 24))
MADRU = set(range(0, 4))


def bloco(nome, horas):
    ch = [c for c, (_d, h) in primeira.items() if h in horas]
    tot_reg = Counter(meta[c][0] for c in primeira)
    reg = Counter(meta[c][0] for c in ch)
    loja = Counter(meta[c][1] for c in ch)
    cat = Counter(meta[c][2] for c in ch)
    perda = sum(perda_ch[c] for c in ch)

    print(f"\n{'='*66}\n{nome} — {len(ch)} chamados ({len(ch)/len(primeira)*100:.1f}% do mes) "
          f"— R$ {perda:,.2f}".replace(",", "."))
    print(f"{'='*66}")
    print(f"{'REGIONAL':<26} {'NOITE':>6} {'TOTAL':>7} {'% DA REGIONAL':>14}")
    for r_, n in reg.most_common():
        t = tot_reg[r_]
        print(f"{r_:<26} {n:>6} {t:>7} {n/t*100:>13.1f}%")

    print(f"\n  Top 8 lojas")
    for i, (l, n) in enumerate(loja.most_common(8), 1):
        print(f"  {i}. {l[:44]:<44} {n:>3}")

    print(f"\n  Top 6 subcategorias")
    for i, (c_, n) in enumerate(cat.most_common(6), 1):
        print(f"  {i}. {c_[:44]:<44} {n:>3}")


print(f"Chamados no mes: {len(primeira)}")
bloco("NOITE — 19h as 23h", NOITE)
bloco("MADRUGADA — 00h as 03h", MADRU)
bloco("NOTURNO COMPLETO — 19h as 03h", NOITE | MADRU)
