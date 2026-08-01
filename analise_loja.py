#!/usr/bin/env python3
"""Top lojas de uma regional por numero de chamados, com o motivo (subcategoria).

Uso: python analise_loja.py <arquivo.xlsx> "SP SUL" [quantas]
"""
import sys
from collections import Counter, defaultdict

import openpyxl

arq = sys.argv[1] if len(sys.argv) > 1 else "detalhada_2607.xlsx"
alvo = sys.argv[2] if len(sys.argv) > 2 else "SP SUL"
topn = int(sys.argv[3]) if len(sys.argv) > 3 else 5

wb = openpyxl.load_workbook(arq, data_only=True)
ws = wb["Export"]

# chamado -> dados; perda rateada quando a linha traz varios chamados
primeira, meta = {}, {}
perda_ch = defaultdict(float)
dias_ch = defaultdict(set)

for r in range(2, ws.max_row + 1):
    regional = ws.cell(row=r, column=3).value
    bkn = ws.cell(row=r, column=4).value
    loja = ws.cell(row=r, column=5).value
    data = ws.cell(row=r, column=6).value
    hora = ws.cell(row=r, column=7).value
    subcat = ws.cell(row=r, column=10).value
    chamados = ws.cell(row=r, column=11).value
    perda = float(ws.cell(row=r, column=12).value or 0)
    if regional != alvo or data is None or hora is None:
        continue
    partes = [c.strip() for c in str(chamados or "").split(",") if c.strip()]
    if not partes:
        continue
    for c in partes:
        perda_ch[c] += perda / len(partes)
        dias_ch[c].add(data.date())
        chave = (data.date(), int(hora))
        if c not in primeira or chave < primeira[c]:
            primeira[c] = chave
            meta[c] = (loja or "", str(bkn or ""), subcat or "")

por_loja = defaultdict(list)
for c in primeira:
    por_loja[meta[c][0]].append(c)

rank = sorted(por_loja.items(), key=lambda x: (-len(x[1]), x[0]))
tot_ch = len(primeira)
tot_perda = sum(perda_ch.values())

print(f"REGIONAL {alvo} — {tot_ch} chamados, R$ {tot_perda:,.2f}, "
      f"{len(por_loja)} lojas com chamado".replace(",", "."))
print(f"Periodo: {min(d for s in dias_ch.values() for d in s)} a "
      f"{max(d for s in dias_ch.values() for d in s)}")
print()

for i, (loja, chs) in enumerate(rank[:topn], 1):
    perda = sum(perda_ch[c] for c in chs)
    bkn = meta[chs[0]][1]
    dias = len({d for c in chs for d in dias_ch[c]})
    horas = Counter(primeira[c][1] for c in chs)
    cats = Counter(meta[c][2] for c in chs)
    print(f"{i}. {loja}  (BKN {bkn})")
    print(f"   {len(chs)} chamados em {dias} dias · R$ {perda:,.2f} "
          f"· {perda/tot_perda*100:.1f}% da regional".replace(",", "."))
    print(f"   motivos:")
    for c_, n in cats.most_common(5):
        pp = sum(perda_ch[c] for c in chs if meta[c][2] == c_)
        print(f"      {n:>3}x  {c_[:40]:<40} R$ {pp:>11,.2f}".replace(",", "."))
    faixa = ", ".join(f"{h:02d}h({n})" for h, n in sorted(horas.items()) if n >= 2)
    print(f"   horarios recorrentes: {faixa or '-'}")
    print()
