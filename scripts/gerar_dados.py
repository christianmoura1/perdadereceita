#!/usr/bin/env python3
"""
Gera data/dados.json e data/detalhe-mes.json a partir do Excel exportado do BI
(Power BI -> Export data, aba "Export": Marca, Diretoria, Regional, BKN, Loja,
Data, Hora, Canal, Motivo, Subcategoria, Chamados, Perda de receita) e dos
percentuais oficiais (%) lidos na tela do BI para cada regional + Brasil Total.

Uso normal (mesmo mês):
    python3 scripts/gerar_dados.py --excel arquivo.xlsx --pct pct.json

Virada de mês (o Excel traz um mês novo): o script arquiva o mês corrente em
`historico` e cria o novo mês. Nesse caso é preciso informar os novos objetivos
(Proj. Venda / Obj. Mês / Obj. Dia / nº restaurantes por regional):
    python3 scripts/gerar_dados.py --excel agosto.xlsx --pct pct.json --objetivos objetivos.json
ou, para repetir os objetivos do mês anterior:
    python3 scripts/gerar_dados.py --excel agosto.xlsx --pct pct.json --manter-objetivos

pct.json — % oficiais da tela do BI (aceita o nome exibido no BI ou a chave
interna; "BRASIL" é obrigatório):
    { "BK É FOGO NORTE": 0.8, ..., "SUL": 0.9, "BRASIL": 0.88 }

objetivos.json — mesmo formato de chaves, com:
    { "RJ": {"rest": 93, "projVenda": 45379559, "objMes": 335809, "objDia": 10833}, ... }

O script:
  1. Usa o data/dados.json atual para montar automaticamente `anterior`
     (snapshot do ciclo anterior, para a seção "Mudanças Importantes").
  2. Lê o Excel e agrega por Regional+Data (R$) e o detalhamento completo
     por ocorrência (loja/hora/canal/categoria/chamado).
  3. Aplica os percentuais oficiais informados.
  4. Atualiza a aba Mensal (regionaisData[reg]['2026'][mes] e brasil2026[mes]).
  5. Valida consistência (somas batendo) antes de gravar.
  6. Grava data/dados.json e data/detalhe-mes.json.
"""
import argparse
import calendar
import json
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent

MES_NOMES = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

LABEL_REG = {"SP CENTRO LITORAL": "BK É FOGO CENTRO LITORAL"}
REVERSE_LABEL = {v: k for k, v in LABEL_REG.items()}


def normaliza_regional(nome):
    """Aceita tanto a chave interna (SP CENTRO LITORAL) quanto o nome exibido
    no BI (BK É FOGO CENTRO LITORAL)."""
    return REVERSE_LABEL.get(nome, nome)


def ler_excel(caminho):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb["Export"]
    linhas = []
    for r in range(2, ws.max_row + 1):
        regional = ws.cell(row=r, column=3).value
        bkn = ws.cell(row=r, column=4).value
        loja = ws.cell(row=r, column=5).value
        data = ws.cell(row=r, column=6).value
        hora = ws.cell(row=r, column=7).value
        canal = ws.cell(row=r, column=8).value
        subcat = ws.cell(row=r, column=10).value
        chamados = ws.cell(row=r, column=11).value
        perda = ws.cell(row=r, column=12).value
        if regional is None or data is None:
            continue
        linhas.append({
            "regional": normaliza_regional(regional),
            "bkn": str(bkn) if bkn is not None else "",
            "loja": loja or "",
            "data": data.date(),
            "hora": int(hora) if hora is not None else None,
            "canal": canal or "",
            "cat": subcat or "",
            "chamado": chamados or "",
            "perda": float(perda) if perda is not None else 0.0,
        })
    return linhas


def build_lojas(linhas):
    agg = {}
    for l in linhas:
        k = (l["bkn"], l["loja"])
        if k not in agg:
            agg[k] = {"bkn": l["bkn"], "nome": l["loja"], "perda": 0.0, "oc": 0}
        agg[k]["perda"] += l["perda"]
        agg[k]["oc"] += 1
    out = list(agg.values())
    out.sort(key=lambda x: -x["perda"])
    return out


def build_cats(linhas):
    agg = {}
    for l in linhas:
        k = l["cat"] or "Outros"
        if k not in agg:
            agg[k] = {"nome": k, "perda": 0.0, "oc": 0}
        agg[k]["perda"] += l["perda"]
        agg[k]["oc"] += 1
    out = list(agg.values())
    out.sort(key=lambda x: -x["perda"])
    return out


def build_horas(linhas):
    agg = defaultdict(float)
    for l in linhas:
        if l["hora"] is not None:
            agg[str(l["hora"])] += l["perda"]
    return dict(agg)


def build_ocs(linhas):
    return [{
        "hora": l["hora"],
        "loja": l["loja"],
        "cat": l["cat"],
        "chamado": l["chamado"],
        "perda": round(l["perda"], 2),
    } for l in linhas]


def build_det_agg(linhas):
    por_regional = defaultdict(list)
    for l in linhas:
        por_regional[l["regional"]].append(l)

    det = {}
    for reg, itens in por_regional.items():
        total_perda = sum(i["perda"] for i in itens)
        datas = sorted(set(i["data"] for i in itens))
        por_dia = {}
        for d in datas:
            do_dia = [i for i in itens if i["data"] == d]
            por_dia[d.isoformat()] = {
                "total": round(sum(i["perda"] for i in do_dia), 2),
                "oc": len(do_dia),
                "cats": build_cats(do_dia),
                "lojas": build_lojas(do_dia),
                "horas": build_horas(do_dia),
                "ocs": build_ocs(do_dia),
            }
        det[reg] = {
            "totalPerda": round(total_perda, 2),
            "oc": len(itens),
            "lojas": build_lojas(itens),
            "cats": build_cats(itens),
            "porDia": por_dia,
            "periodo": {"min": datas[0].isoformat(), "max": datas[-1].isoformat()},
        }
    return det


def snapshot_anterior(mes_atual, reg_names):
    """Monta o snapshot do ciclo anterior a partir do estado atual do JSON
    (equivalente à antiga 'Fase 0' manual do prompt mestre)."""
    dias_no_mes = mes_atual["diasNoMes"]
    total_acum = sum(mes_atual["extra"][r]["acum"] for r in reg_names)
    dias_com_dados = len([v for v in mes_atual["brasilVals"] if v is not None])
    if dias_com_dados > 0:
        media = total_acum / dias_com_dados
        proj = total_acum + media * (dias_no_mes - dias_com_dados)
    else:
        proj = 0
    return {
        "pctTotal": mes_atual["pctTotal"],
        "regionais": {r: mes_atual["extra"][r]["pctMes"] for r in reg_names},
        "projFechamento": proj,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True)
    ap.add_argument("--pct", required=True, help="JSON com os % oficiais da tela do BI")
    ap.add_argument("--objetivos", help="JSON com rest/projVenda/objMes/objDia por regional (virada de mês)")
    ap.add_argument("--manter-objetivos", action="store_true",
                    help="Na virada de mês, repetir os objetivos do mês anterior")
    ap.add_argument("--dados-dir", default=None, help="Diretório data/ alternativo (para testes)")
    ap.add_argument("--dry-run", action="store_true", help="Não grava arquivos, só valida e mostra o resumo")
    args = ap.parse_args()

    data_dir = Path(args.dados_dir) if args.dados_dir else (ROOT / "data")
    dados_path = data_dir / "dados.json"
    detalhe_path = data_dir / "detalhe-mes.json"

    with open(dados_path) as f:
        dados = json.load(f)
    with open(args.pct) as f:
        pct_raw = json.load(f)

    pct = {normaliza_regional(k): v for k, v in pct_raw.items() if k != "BRASIL"}
    if "BRASIL" not in pct_raw or pct_raw["BRASIL"] is None:
        sys.exit("pct.json precisa da chave BRASIL com o % oficial do Brasil Total")
    pct_brasil = pct_raw["BRASIL"]

    linhas = ler_excel(args.excel)
    if not linhas:
        sys.exit("Excel sem linhas válidas")

    datas_presentes = sorted(set(l["data"] for l in linhas))
    ano, mes = datas_presentes[-1].year, datas_presentes[-1].month
    for d in datas_presentes:
        if (d.year, d.month) != (ano, mes):
            sys.exit(f"Excel contém datas de mais de um mês: {d} vs {ano}-{mes:02d}")
    ultimo_dia = datas_presentes[-1].day
    dias_no_mes = calendar.monthrange(ano, mes)[1]

    mes_atual = dados["mesAtual"]
    reg_names = list(mes_atual["extra"].keys())
    for r in reg_names:
        if r not in pct or pct[r] is None:
            sys.exit(f"Falta % oficial para regional: {r}")

    virada = (ano, mes) != (mes_atual["ano"], mes_atual["mes"])

    if virada:
        if (ano, mes) < (mes_atual["ano"], mes_atual["mes"]):
            sys.exit(f"Excel traz mês anterior ao corrente ({ano}-{mes:02d}); nada a fazer")
        # objetivos do novo mês
        if args.objetivos:
            with open(args.objetivos) as f:
                obj_raw = json.load(f)
            objetivos = {normaliza_regional(k): v for k, v in obj_raw.items()}
            for r in reg_names:
                if r not in objetivos:
                    sys.exit(f"objetivos.json sem a regional: {r}")
                for campo in ("rest", "projVenda", "objMes", "objDia"):
                    if campo not in objetivos[r] or objetivos[r][campo] is None:
                        sys.exit(f"objetivos.json: falta '{campo}' em {r}")
        elif args.manter_objetivos:
            objetivos = {r: {k: mes_atual["extra"][r][k] for k in ("rest", "projVenda", "objMes", "objDia")}
                         for r in reg_names}
            print("AVISO: repetindo objetivos do mês anterior (--manter-objetivos)")
        else:
            sys.exit(
                f"O Excel traz {MES_NOMES[mes-1]}/{ano}, mas o mês corrente é "
                f"{mes_atual['nome']}/{mes_atual['ano']} — isso é uma virada de mês.\n"
                "Informe --objetivos objetivos.json (novos Proj. Venda / Obj. Mês / Obj. Dia / Rest.)\n"
                "ou --manter-objetivos para repetir os do mês anterior."
            )
        # arquiva o mês corrente
        arquivado = {k: v for k, v in mes_atual.items() if k != "anterior"}
        dados["historico"].insert(0, arquivado)
        anterior = {
            "pctTotal": mes_atual["pctTotal"],
            "regionais": {r: mes_atual["extra"][r]["pctMes"] for r in reg_names},
            "projFechamento": sum(mes_atual["extra"][r]["acum"] for r in reg_names),
        }
        print(f"Virada de mês: {mes_atual['nome']}/{mes_atual['ano']} arquivado; iniciando {MES_NOMES[mes-1]}/{ano}")
    else:
        objetivos = {r: {k: mes_atual["extra"][r][k] for k in ("rest", "projVenda", "objMes", "objDia")}
                     for r in reg_names}
        anterior = snapshot_anterior(mes_atual, reg_names)

    # ---- Agregação por Regional + Data ----
    por_regional_data = defaultdict(dict)
    for l in linhas:
        por_regional_data[l["regional"]].setdefault(l["data"], 0.0)
        por_regional_data[l["regional"]][l["data"]] += l["perda"]

    novo_data = {}
    novo_extra = {}
    for reg in reg_names:
        valores_dia = por_regional_data.get(reg, {})
        arr = []
        for dia in range(1, dias_no_mes + 1):
            if dia <= ultimo_dia:
                arr.append(round(valores_dia.get(date(ano, mes, dia), 0.0), 2))
            else:
                arr.append(None)
        novo_data[reg] = arr

        acum = round(sum(v for v in arr if v is not None), 2)
        dia_ult = arr[ultimo_dia - 1] or 0.0
        obj = objetivos[reg]
        novo_extra[reg] = {
            "rest": obj["rest"],
            "projVenda": obj["projVenda"],
            "objMes": obj["objMes"],
            "objDia": obj["objDia"],
            "diaUlt": round(dia_ult, 2),
            "acumAntes": round(acum - dia_ult, 2),
            "acum": acum,
            "perdaLoja": round(acum / obj["rest"], 2) if obj["rest"] else 0,
            "pctMes": pct[reg],
        }

    brasil_vals = []
    for dia in range(1, dias_no_mes + 1):
        if dia <= ultimo_dia:
            brasil_vals.append(round(sum((novo_data[reg][dia - 1] or 0.0) for reg in reg_names), 2))
        else:
            brasil_vals.append(None)

    # ---- Validações ----
    soma_acum = round(sum(novo_extra[r]["acum"] for r in reg_names), 2)
    soma_brasil = round(sum(v for v in brasil_vals if v is not None), 2)
    if abs(soma_acum - soma_brasil) > 1:
        sys.exit(f"Inconsistência: soma extra.acum ({soma_acum}) != soma brasilVals ({soma_brasil})")

    # ---- Atualiza aba Mensal ----
    mes_idx = mes - 1
    for reg in reg_names:
        dados["regionaisData"][reg]["2026"][mes_idx] = novo_extra[reg]["acum"]
    dados["brasil2026"][mes_idx] = soma_acum

    dados["mesAtual"] = {
        "ano": ano,
        "mes": mes,
        "nome": MES_NOMES[mes - 1],
        "diasNoMes": dias_no_mes,
        "metaDiaBrasil": sum(objetivos[r]["objDia"] for r in reg_names),
        "pctTotal": pct_brasil,
        "ultimoDia": ultimo_dia,
        "atualizadoEm": datas_presentes[-1].isoformat(),
        "data": novo_data,
        "extra": novo_extra,
        "brasilVals": brasil_vals,
        "anterior": anterior,
    }

    det_agg = build_det_agg(linhas)

    print(f"Mês: {MES_NOMES[mes-1]}/{ano} — último dia com dados: {ultimo_dia}")
    print(f"Total Brasil acumulado: R$ {soma_acum:,.2f}")
    print(f"% oficial Brasil: {pct_brasil}%")
    for reg in reg_names:
        e = novo_extra[reg]
        print(f"  {reg:22s} acum=R$ {e['acum']:>12,.2f}  pct={e['pctMes']}%  diaUlt=R$ {e['diaUlt']:>10,.2f}")

    if args.dry_run:
        print("\n--dry-run: nada foi gravado.")
        return

    data_dir.mkdir(parents=True, exist_ok=True)
    with open(dados_path, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, separators=(",", ":"))
    with open(detalhe_path, "w", encoding="utf-8") as f:
        json.dump(det_agg, f, ensure_ascii=False, separators=(",", ":"))
    print(f"\nGravado: {dados_path}")
    print(f"Gravado: {detalhe_path}")


if __name__ == "__main__":
    main()
