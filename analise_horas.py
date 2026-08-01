#!/usr/bin/env python3
"""Distribuicao horaria dos chamados, a partir do export detalhado do BI."""
import sys
from collections import Counter, defaultdict

import openpyxl

arq = sys.argv[1] if len(sys.argv) > 1 else "detalhada_2607.xlsx"
wb = openpyxl.load_workbook(arq, data_only=True)
ws = wb["Export"]

primeira = {}          # chamado -> (data, hora) minimo
linhas_por_hora = Counter()
perda_por_hora = defaultdict(float)
reg_do_chamado = {}
linhas = 0

for r in range(2, ws.max_row + 1):
    regional = ws.cell(row=r, column=3).value
    data = ws.cell(row=r, column=6).value
    hora = ws.cell(row=r, column=7).value
    chamados = ws.cell(row=r, column=11).value
    perda = ws.cell(row=r, column=12).value
    if regional is None or data is None or hora is None:
        continue
    linhas += 1
    h = int(hora)
    linhas_por_hora[h] += 1
    perda_por_hora[h] += float(perda or 0)
    for c in str(chamados or "").split(","):
        c = c.strip()
        if not c:
            continue
        chave = (data.date(), h)
        if c not in primeira or chave < primeira[c]:
            primeira[c] = chave
            reg_do_chamado[c] = regional

aberturas = Counter(h for _d, h in primeira.values())
total_ch = len(primeira)

print(f"Linhas validas ......... {linhas:,}".replace(",", "."))
print(f"Chamados distintos ..... {total_ch:,}".replace(",", "."))
print()
print("HORA | ABERTURAS  %     | LINHAS(h de perda) | PERDA R$")
print("-" * 62)
for h in range(24):
    ab = aberturas.get(h, 0)
    if ab == 0 and linhas_por_hora.get(h, 0) == 0:
        continue
    pctm = ab / total_ch * 100 if total_ch else 0
    barra = "#" * round(pctm / 1.2)
    print(f" {h:02d}h | {ab:>6}  {pctm:5.1f}% | {linhas_por_hora.get(h,0):>10} "
          f"| {perda_por_hora.get(h,0):>12,.2f}  {barra}".replace(",", "."))

print()
print("TOP 6 HORARIOS POR ABERTURA")
for i, (h, n) in enumerate(aberturas.most_common(6), 1):
    print(f"  {i}. {h:02d}h — {n} chamados ({n/total_ch*100:.1f}%)")

print()
print("TOP 6 HORARIOS POR PERDA R$")
for i, (h, v) in enumerate(sorted(perda_por_hora.items(), key=lambda x: -x[1])[:6], 1):
    print(f"  {i}. {h:02d}h — R$ {v:,.2f}".replace(",", "."))
